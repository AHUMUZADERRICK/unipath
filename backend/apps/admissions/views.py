from decimal import Decimal
import secrets
import csv
import json
from io import BytesIO
from datetime import datetime

from django.contrib.auth.hashers import check_password, make_password
from django.contrib.auth import authenticate
from django.utils import timezone
from django.db.models import Count, Avg, Q
from django.http import HttpResponse
from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib import colors
from rest_framework import status
from rest_framework.exceptions import AuthenticationFailed, PermissionDenied
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.response import Response
from rest_framework.views import APIView

from .models import (
    AdminSessionToken,
    Candidate,
    Course,
    CourseSubjectRequirement,
    CourseCutoffHistory,
    PlannerCourse,
    StudentResult,
    StudentSubmission,
    StudentStrengthProfile,
    Subject,
    UCEGrade,
    University,
)
from .serializers import (
    AdminCandidateUpdateSerializer,
    AdminCutoffExcelUploadSerializer,
    AdminCourseCutoffUpdateSerializer,
    AdminLoginSerializer,
    AnalyticsSerializer,
    CalculateWeightInputSerializer,
    CandidateLoginSerializer,
    CandidateSignupSerializer,
    CourseSubjectRequirementSerializer,
    CourseSubjectRequirementUpdateSerializer,
    PlannerCourseCreateUpdateSerializer,
    PlannerCourseSerializer,
)
from .services import evaluate_courses, get_eligible_evaluations, profile_student_strengths


def _serialise_candidate(candidate: Candidate) -> dict:
    return {
        "id": candidate.id,
        "first_name": candidate.first_name,
        "last_name": candidate.last_name,
        "email": candidate.email,
        "whatsapp_number": candidate.whatsapp_number,
        "index_number": candidate.index_number,
        "gender": candidate.gender,
    }


class CandidateSignupAPIView(APIView):
    serializer_class = CandidateSignupSerializer

    def post(self, request, *args, **kwargs):
        serializer = self.serializer_class(data=request.data)
        serializer.is_valid(raise_exception=True)

        payload = serializer.validated_data
        email = payload["email"].strip().lower()

        if Candidate.objects.filter(email=email).exists():
            return Response(
                {"detail": "An account with this email already exists."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        candidate, created = Candidate.objects.get_or_create(
            whatsapp_number=payload["whatsapp_number"].strip(),
            index_number=payload["index_number"].strip().upper(),
            defaults={
                "first_name": payload["first_name"].strip(),
                "last_name": payload["last_name"].strip(),
                "email": email,
                "password_hash": make_password(payload["password"]),
                "gender": payload["gender"],
            },
        )

        if not created:
            candidate.first_name = payload["first_name"].strip()
            candidate.last_name = payload["last_name"].strip()
            candidate.email = email
            candidate.password_hash = make_password(payload["password"])
            candidate.gender = payload["gender"]
            candidate.save(
                update_fields=["first_name", "last_name", "email", "password_hash", "gender"]
            )

        return Response(
            {
                "candidate": _serialise_candidate(candidate),
            },
            status=status.HTTP_201_CREATED if created else status.HTTP_200_OK,
        )


class CandidateLoginAPIView(APIView):
    serializer_class = CandidateLoginSerializer

    def post(self, request, *args, **kwargs):
        serializer = self.serializer_class(data=request.data)
        serializer.is_valid(raise_exception=True)

        payload = serializer.validated_data
        candidate = Candidate.objects.filter(email=payload["email"].strip().lower()).first()

        if not candidate or not check_password(payload["password"], candidate.password_hash):
            return Response(
                {"detail": "Invalid email or password."},
                status=status.HTTP_401_UNAUTHORIZED,
            )

        return Response({"candidate": _serialise_candidate(candidate)}, status=status.HTTP_200_OK)


def _extract_bearer_token(request) -> str | None:
    auth_header = request.headers.get("Authorization", "")
    if not auth_header.startswith("Bearer "):
        return None
    return auth_header.replace("Bearer ", "", 1).strip()


class AdminTokenProtectedAPIView(APIView):
    def _get_admin_user_and_session(self, request):
        token = _extract_bearer_token(request)
        if not token:
            raise AuthenticationFailed("Missing admin token.")

        session = AdminSessionToken.objects.select_related("user").filter(token=token).first()
        if not session:
            raise AuthenticationFailed("Invalid or expired admin token.")

        if not session.user.is_staff:
            raise PermissionDenied("Admin access required.")

        session.last_used_at = timezone.now()
        session.save(update_fields=["last_used_at"])
        return session.user, session


class AdminLoginAPIView(APIView):
    serializer_class = AdminLoginSerializer

    def post(self, request, *args, **kwargs):
        serializer = self.serializer_class(data=request.data)
        serializer.is_valid(raise_exception=True)

        username = serializer.validated_data["username"].strip()
        password = serializer.validated_data["password"]

        user = authenticate(username=username, password=password)
        if not user or not user.is_staff:
            return Response({"detail": "Invalid admin credentials."}, status=status.HTTP_401_UNAUTHORIZED)

        token_value = secrets.token_urlsafe(40)
        session = AdminSessionToken.objects.create(user=user, token=token_value)
        return Response(
            {
                "token": session.token,
                "admin": {
                    "id": user.id,
                    "username": user.username,
                    "email": user.email,
                    "is_superuser": user.is_superuser,
                },
            },
            status=status.HTTP_200_OK,
        )


class AdminLogoutAPIView(AdminTokenProtectedAPIView):
    def post(self, request, *args, **kwargs):
        _, session = self._get_admin_user_and_session(request)
        session.delete()
        return Response({"detail": "Logged out."}, status=status.HTTP_200_OK)


class AdminDashboardAPIView(AdminTokenProtectedAPIView):
    def get(self, request, *args, **kwargs):
        _, _ = self._get_admin_user_and_session(request)

        today = timezone.localdate()
        dashboard = {
            "total_candidates": Candidate.objects.count(),
            "total_submissions": StudentSubmission.objects.count(),
            "submissions_today": StudentSubmission.objects.filter(created_at__date=today).count(),
            "total_courses": Course.objects.count(),
            "total_universities": University.objects.count(),
            "total_admin_sessions": AdminSessionToken.objects.count(),
            "gender_distribution": list(
                Candidate.objects.values("gender").annotate(total=Count("id")).order_by("gender")
            ),
            "recent_submissions": list(
                StudentSubmission.objects.select_related("candidate")
                .order_by("-created_at")[:8]
                .values("id", "gender", "created_at", "candidate__email", "candidate__index_number")
            ),
        }
        return Response(dashboard, status=status.HTTP_200_OK)


class AdminCandidatesAPIView(AdminTokenProtectedAPIView):
    def get(self, request, *args, **kwargs):
        _, _ = self._get_admin_user_and_session(request)
        candidates = list(
            Candidate.objects.order_by("-created_at").values(
                "id",
                "first_name",
                "last_name",
                "email",
                "whatsapp_number",
                "index_number",
                "gender",
                "created_at",
            )[:200]
        )
        return Response({"candidates": candidates}, status=status.HTTP_200_OK)


class AdminCandidateDetailAPIView(AdminTokenProtectedAPIView):
    serializer_class = AdminCandidateUpdateSerializer

    def patch(self, request, candidate_id, *args, **kwargs):
        _, _ = self._get_admin_user_and_session(request)
        candidate = Candidate.objects.filter(id=candidate_id).first()
        if not candidate:
            return Response({"detail": "Candidate not found."}, status=status.HTTP_404_NOT_FOUND)

        serializer = self.serializer_class(data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        payload = serializer.validated_data

        for field in ["first_name", "last_name", "email", "whatsapp_number", "index_number", "gender"]:
            if field in payload:
                setattr(candidate, field, payload[field])

        if "index_number" in payload:
            candidate.index_number = candidate.index_number.strip().upper()
        if "email" in payload:
            candidate.email = candidate.email.strip().lower()

        candidate.save()
        return Response({"candidate": _serialise_candidate(candidate)}, status=status.HTTP_200_OK)


class AdminCoursesAPIView(AdminTokenProtectedAPIView):
    def get(self, request, *args, **kwargs):
        _, _ = self._get_admin_user_and_session(request)
        courses = list(
            Course.objects.select_related("university")
            .order_by("name")
            .values("id", "name", "cutoff_weight", "duration", "university__name")
        )
        return Response({"courses": courses}, status=status.HTTP_200_OK)


class AdminCourseCutoffAPIView(AdminTokenProtectedAPIView):
    serializer_class = AdminCourseCutoffUpdateSerializer

    def patch(self, request, course_id, *args, **kwargs):
        _, _ = self._get_admin_user_and_session(request)
        course = Course.objects.filter(id=course_id).first()
        if not course:
            return Response({"detail": "Course not found."}, status=status.HTTP_404_NOT_FOUND)

        serializer = self.serializer_class(data=request.data)
        serializer.is_valid(raise_exception=True)
        payload = serializer.validated_data

        course.cutoff_weight = payload["cutoff_weight"]
        course.save(update_fields=["cutoff_weight"])

        year = payload.get("year")
        if year:
            CourseCutoffHistory.objects.update_or_create(
                course=course,
                year=year,
                defaults={"cutoff_weight": payload["cutoff_weight"]},
            )

        return Response(
            {
                "course": {
                    "id": course.id,
                    "name": course.name,
                    "cutoff_weight": course.cutoff_weight,
                }
            },
            status=status.HTTP_200_OK,
        )


class AdminCutoffTemplateDownloadAPIView(AdminTokenProtectedAPIView):
    def get(self, request, *args, **kwargs):
        _, _ = self._get_admin_user_and_session(request)

        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = "attachment; filename=course_cutoff_template.csv"

        writer = csv.writer(response)
        writer.writerow(["university", "course", "cutoff_weight", "year", "duration"])
        writer.writerow(["Makerere University", "Computer Science", "43.00", "2026", "3"])
        writer.writerow(["Kyambogo University", "Information Technology", "41.00", "2026", "3"])
        return response


class AdminCutoffHistoryTemplateDownloadAPIView(AdminTokenProtectedAPIView):
    def get(self, request, *args, **kwargs):
        _, _ = self._get_admin_user_and_session(request)

        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = "attachment; filename=course_cutoff_history_template.csv"

        writer = csv.writer(response)
        writer.writerow(["university", "course", "year", "cutoff_weight"])
        writer.writerow(["Makerere University", "Computer Science", "2026", "43.00"])
        writer.writerow(["Kyambogo University", "Information Technology", "2026", "41.00"])
        return response


class AdminCutoffExcelUploadAPIView(AdminTokenProtectedAPIView):
    serializer_class = AdminCutoffExcelUploadSerializer
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request, *args, **kwargs):
        _, _ = self._get_admin_user_and_session(request)

        serializer = self.serializer_class(data=request.data)
        serializer.is_valid(raise_exception=True)
        upload = serializer.validated_data["file"]

        filename = (upload.name or "").lower()
        if not (filename.endswith(".xlsx") or filename.endswith(".xlsm")):
            return Response(
                {"detail": "Only Excel .xlsx/.xlsm files are supported."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            workbook = load_workbook(filename=BytesIO(upload.read()), data_only=True)
        except Exception:
            return Response(
                {"detail": "Could not read Excel file. Ensure it is a valid .xlsx workbook."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            return Response({"detail": "Excel file is empty."}, status=status.HTTP_400_BAD_REQUEST)

        header_cells = [str(cell).strip().lower() if cell is not None else "" for cell in rows[0]]
        required_columns = ["university", "course", "cutoff_weight"]
        missing_columns = [column for column in required_columns if column not in header_cells]
        if missing_columns:
            return Response(
                {"detail": f"Missing required columns: {', '.join(missing_columns)}."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        header_index = {name: index for index, name in enumerate(header_cells)}
        created_count = 0
        updated_count = 0
        history_rows = 0
        skipped_rows = []

        for excel_row_number, row_values in enumerate(rows[1:], start=2):
            if not row_values or all(value in (None, "") for value in row_values):
                continue

            def value_of(column_name):
                index = header_index.get(column_name)
                if index is None or index >= len(row_values):
                    return ""
                value = row_values[index]
                return "" if value is None else str(value).strip()

            university_name = value_of("university")
            course_name = value_of("course")
            cutoff_raw = value_of("cutoff_weight")
            year_raw = value_of("year")
            duration_raw = value_of("duration")

            if not university_name or not course_name or not cutoff_raw:
                skipped_rows.append(
                    {
                        "row": excel_row_number,
                        "reason": "Missing university/course/cutoff_weight.",
                    }
                )
                continue

            try:
                cutoff_weight = Decimal(cutoff_raw)
            except Exception:
                skipped_rows.append(
                    {
                        "row": excel_row_number,
                        "reason": f"Invalid cutoff_weight '{cutoff_raw}'.",
                    }
                )
                continue

            duration = 3
            if duration_raw:
                try:
                    duration = int(float(duration_raw))
                except Exception:
                    skipped_rows.append(
                        {
                            "row": excel_row_number,
                            "reason": f"Invalid duration '{duration_raw}'.",
                        }
                    )
                    continue

            university, _ = University.objects.get_or_create(
                name=university_name,
                defaults={"location": "Uganda"},
            )
            course, created = Course.objects.update_or_create(
                name=course_name,
                university=university,
                defaults={"cutoff_weight": cutoff_weight, "duration": duration},
            )

            if created:
                created_count += 1
            else:
                updated_count += 1

            if year_raw:
                try:
                    year = int(float(year_raw))
                except Exception:
                    skipped_rows.append(
                        {
                            "row": excel_row_number,
                            "reason": f"Invalid year '{year_raw}'.",
                        }
                    )
                else:
                    CourseCutoffHistory.objects.update_or_create(
                        course=course,
                        year=year,
                        defaults={"cutoff_weight": cutoff_weight},
                    )
                    history_rows += 1

        return Response(
            {
                "created_courses": created_count,
                "updated_courses": updated_count,
                "updated_history_rows": history_rows,
                "skipped_rows": skipped_rows,
            },
            status=status.HTTP_200_OK,
        )


class CalculateWeightAPIView(APIView):
    serializer_class = CalculateWeightInputSerializer

    def post(self, request, *args, **kwargs):
        serializer = self.serializer_class(data=request.data)
        serializer.is_valid(raise_exception=True)

        payload = serializer.validated_data
        candidate_id = payload.get("candidate_id")
        candidate = Candidate.objects.filter(id=candidate_id).first() if candidate_id else None
        gender = candidate.gender if candidate else payload["gender"]
        uace_results = payload["uace_results"]
        uce_grades = payload["uce_grades"]

        submission = StudentSubmission.objects.create(gender=gender, candidate=candidate)
        self._persist_student_results(submission, uace_results, uce_grades)

        evaluations = evaluate_courses(
            student_results=uace_results,
            uce_grades=uce_grades,
            gender=gender,
        )

        eligible_courses = [
            {
                "course": evaluation.course_name,
                "university": evaluation.university_name,
                "cutoff": evaluation.cutoff_weight,
                "calculated_weight": evaluation.calculated_weight,
                "is_borderline": evaluation.is_borderline,
                "recommendation_band": evaluation.recommendation_band,
                "explanation": evaluation.explanation,
            }
            for evaluation in get_eligible_evaluations(evaluations)
        ]

        borderline_courses = [
            {
                "course": evaluation.course_name,
                "university": evaluation.university_name,
                "cutoff": evaluation.cutoff_weight,
                "calculated_weight": evaluation.calculated_weight,
                "recommendation_band": evaluation.recommendation_band,
                "explanation": evaluation.explanation,
            }
            for evaluation in evaluations
            if evaluation.is_borderline and not evaluation.is_eligible
        ]

        course_evaluations = [
            {
                "course": evaluation.course_name,
                "university": evaluation.university_name,
                "cutoff": evaluation.cutoff_weight,
                "calculated_weight": evaluation.calculated_weight,
                "is_eligible": evaluation.is_eligible,
                "is_borderline": evaluation.is_borderline,
                "recommendation_band": evaluation.recommendation_band,
                "missing_essential_subjects": evaluation.missing_essential_subjects,
                "explanation": evaluation.explanation,
            }
            for evaluation in evaluations
        ]

        recommendation_groups = {
            "top_chances": [
                item
                for item in course_evaluations
                if item["recommendation_band"] == "top_chance"
            ],
            "safe_options": [
                item
                for item in course_evaluations
                if item["recommendation_band"] == "safe_option"
            ],
            "ambitious_choices": [
                item
                for item in course_evaluations
                if item["recommendation_band"] == "ambitious_choice"
            ],
        }

        final_weight = max((item.calculated_weight for item in evaluations), default=Decimal("0.00"))

        # Generate strength profile
        strength_profile = profile_student_strengths(serializer.validated_data["uace_results"], uce_grades)
        strength_profile_data = {
            "strong_subjects": [
                {
                    "subject_name": s.subject_name,
                    "grade": s.grade,
                    "strength_level": s.strength_level,
                }
                for s in strength_profile.strong_subjects
            ],
            "recommended_course_ids": strength_profile.recommended_course_ids,
            "insights": strength_profile.insights,
        }

        return Response(
            {
                "final_weight": final_weight,
                "eligible_courses": eligible_courses,
                "borderline_courses": borderline_courses,
                "course_evaluations": course_evaluations,
                "recommendation_groups": recommendation_groups,
                "strength_profile": strength_profile_data,
            },
            status=status.HTTP_200_OK,
        )

    @staticmethod
    def _persist_student_results(
        submission: StudentSubmission, uace_results: dict[str, str], uce_grades: list[str]
    ) -> None:
        for subject_name, grade in uace_results.items():
            subject, _ = Subject.objects.get_or_create(name=subject_name.strip())
            StudentResult.objects.create(submission=submission, subject=subject, grade=grade)

        for grade in uce_grades:
            UCEGrade.objects.create(submission=submission, grade=grade)


class UACESubjectListAPIView(APIView):
    def get(self, request, *args, **kwargs):
        subjects = list(Subject.objects.order_by("name").values_list("name", flat=True))
        return Response({"subjects": subjects}, status=status.HTTP_200_OK)


class PlannerCourseListAPIView(APIView):
    """List and save candidate's planned courses."""
    
    def get(self, request, *args, **kwargs):
        candidate_id = request.query_params.get("candidate_id")
        if not candidate_id:
            raise PermissionDenied("candidate_id required")
        
        try:
            candidate = Candidate.objects.get(id=candidate_id)
        except Candidate.DoesNotExist:
            raise PermissionDenied("Candidate not found")
        
        planned = PlannerCourse.objects.filter(candidate=candidate).select_related("course__university")
        courses_data = [
            {
                "id": p.id,
                "course_id": p.course.id,
                "course_name": p.course.name,
                "university": p.course.university.name,
                "rank": p.rank,
                "notes": p.notes,
                "created_at": p.created_at,
                "updated_at": p.updated_at,
            }
            for p in planned.order_by("rank")
        ]
        return Response(courses_data, status=status.HTTP_200_OK)
    
    def post(self, request, *args, **kwargs):
        candidate_id = request.data.get("candidate_id")
        course_id = request.data.get("course_id")
        rank = request.data.get("rank")
        notes = request.data.get("notes", "")
        
        if not candidate_id or not course_id or rank is None:
            raise PermissionDenied("candidate_id, course_id, and rank required")
        
        try:
            candidate = Candidate.objects.get(id=candidate_id)
            course = Course.objects.get(id=course_id)
        except (Candidate.DoesNotExist, Course.DoesNotExist):
            raise PermissionDenied("Candidate or course not found")
        
        # Adjust ranks for existing courses if necessary
        existing = PlannerCourse.objects.filter(candidate=candidate, rank__gte=rank).exclude(course=course)
        for item in existing:
            item.rank += 1
            item.save()
        
        planned, created = PlannerCourse.objects.update_or_create(
            candidate=candidate,
            course=course,
            defaults={"rank": rank, "notes": notes}
        )
        
        return Response(
            {
                "id": planned.id,
                "course_id": planned.course.id,
                "rank": planned.rank,
                "notes": planned.notes,
            },
            status=status.HTTP_201_CREATED if created else status.HTTP_200_OK,
        )
    
    def delete(self, request, *args, **kwargs):
        planner_id = request.query_params.get("planner_id")
        if not planner_id:
            raise PermissionDenied("planner_id required")
        
        try:
            planned = PlannerCourse.objects.get(id=planner_id)
            planned.delete()
            return Response({"message": "Deleted"}, status=status.HTTP_204_NO_CONTENT)
        except PlannerCourse.DoesNotExist:
            raise PermissionDenied("Planned course not found")


class StrengthProfileAPIView(APIView):
    """Get student's strength profile and save it."""
    
    def post(self, request, *args, **kwargs):
        submission_id = request.data.get("submission_id")
        uace_results = request.data.get("uace_results", {})
        uce_grades = request.data.get("uce_grades", [])
        
        if not submission_id or not uace_results:
            raise PermissionDenied("submission_id and uace_results required")
        
        try:
            submission = StudentSubmission.objects.get(id=submission_id)
        except StudentSubmission.DoesNotExist:
            raise PermissionDenied("Submission not found")
        
        # Generate strength profile
        profile = profile_student_strengths(uace_results, uce_grades)
        
        # Store profile
        strong_subjects_data = [
            {
                "subject_name": s.subject_name,
                "grade": s.grade,
                "strength_level": s.strength_level,
            }
            for s in profile.strong_subjects
        ]
        
        strength_profile, created = StudentStrengthProfile.objects.update_or_create(
            submission=submission,
            defaults={
                "strong_subjects": strong_subjects_data,
                "recommended_course_ids": profile.recommended_course_ids,
                "insights": profile.insights,
            }
        )
        
        return Response(
            {
                "strong_subjects": strong_subjects_data,
                "recommended_course_ids": profile.recommended_course_ids,
                "insights": profile.insights,
            },
            status=status.HTTP_201_CREATED if created else status.HTTP_200_OK,
        )
    
    def get(self, request, *args, **kwargs):
        submission_id = request.query_params.get("submission_id")
        if not submission_id:
            raise PermissionDenied("submission_id required")
        
        try:
            profile = StudentStrengthProfile.objects.get(submission_id=submission_id)
            return Response(
                {
                    "strong_subjects": profile.strong_subjects,
                    "recommended_course_ids": profile.recommended_course_ids,
                    "insights": profile.insights,
                },
                status=status.HTTP_200_OK,
            )
        except StudentStrengthProfile.DoesNotExist:
            return Response(
                {"message": "No strength profile yet"},
                status=status.HTTP_404_NOT_FOUND,
            )


class ExportPDFAPIView(APIView):
    """Export eligibility summary as PDF."""
    
    def post(self, request, *args, **kwargs):
        candidate_id = request.data.get("candidate_id")
        evaluations_data = request.data.get("evaluations", [])
        final_weight = request.data.get("final_weight", "0.00")
        
        if not candidate_id or not evaluations_data:
            raise PermissionDenied("candidate_id and evaluations required")
        
        try:
            candidate = Candidate.objects.get(id=candidate_id)
        except Candidate.DoesNotExist:
            raise PermissionDenied("Candidate not found")
        
        # Create PDF
        pdf_buffer = BytesIO()
        doc = SimpleDocTemplate(pdf_buffer, pagesize=letter)
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            textColor=colors.HexColor('#1e40af'),
            spaceAfter=12,
        )
        story.append(Paragraph("Course Eligibility Summary", title_style))
        story.append(Spacer(1, 0.2 * inch))
        
        # Candidate info
        info_style = styles['Normal']
        story.append(Paragraph(f"<b>Candidate:</b> {candidate.first_name} {candidate.last_name}", info_style))
        story.append(Paragraph(f"<b>Email:</b> {candidate.email}", info_style))
        story.append(Paragraph(f"<b>Final Weight:</b> {final_weight}", info_style))
        story.append(Paragraph(f"<b>Generated:</b> {datetime.now().strftime('%Y-%m-%d %H:%M')}", info_style))
        story.append(Spacer(1, 0.3 * inch))
        
        # Eligible courses table
        if evaluations_data:
            story.append(Paragraph("<b>Eligible Courses</b>", styles['Heading2']))
            table_data = [["Course", "University", "Cutoff", "Status"]]
            for eval_item in evaluations_data:
                if eval_item.get("is_eligible"):
                    table_data.append([
                        eval_item.get("course", ""),
                        eval_item.get("university", ""),
                        str(eval_item.get("cutoff", "")),
                        eval_item.get("recommendation_band", ""),
                    ])
            
            if len(table_data) > 1:
                table = Table(table_data)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ]))
                story.append(table)
        
        # Build PDF
        doc.build(story)
        pdf_buffer.seek(0)
        
        filename = f"eligibility_{candidate_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        response = HttpResponse(pdf_buffer.getvalue(), content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


class AnalyticsAPIView(APIView):
    """Get admin analytics on most selected courses and common profiles."""
    
    @staticmethod
    def _require_admin_token(request):
        """Verify admin token."""
        auth_header = request.META.get("HTTP_AUTHORIZATION", "")
        if not auth_header.startswith("Bearer "):
            raise AuthenticationFailed("admin token required")
        
        token = auth_header[7:]
        try:
            session = AdminSessionToken.objects.get(token=token)
            session.last_used_at = timezone.now()
            session.save()
            return session.user
        except AdminSessionToken.DoesNotExist:
            raise AuthenticationFailed("invalid token")
    
    def get(self, request, *args, **kwargs):
        self._require_admin_token(request)
        
        # Most selected courses
        top_courses_data = [
            {
                "course_id": item["course_id"],
                "course_name": item["course__name"],
                "university": item["course__university__name"],
                "count": item["count"],
            }
            for item in PlannerCourse.objects.values(
                "course_id",
                "course__name",
                "course__university__name",
            )
            .annotate(count=Count("id"))
            .order_by("-count")[:10]
        ]
        
        # Common profiles: most frequent gender + top subjects
        common_profiles = []
        gender_counts = StudentSubmission.objects.values("gender").annotate(count=Count("id")).order_by("-count")
        for entry in gender_counts[:5]:
            common_profiles.append({
                "profile_type": "Gender",
                "value": entry["gender"].title(),
                "count": entry["count"],
            })
        
        # Total stats
        total_submissions = StudentSubmission.objects.count()
        total_candidates = Candidate.objects.count()
        avg_weight = Decimal("0.00")
        
        # Calculate average weight from StudentResult data
        from .services import calculate_weight
        results = StudentResult.objects.values("submission").distinct()
        if results:
            weights = []
            for result in results:
                submission = StudentSubmission.objects.get(id=result["submission"])
                uace_results = {r.subject.name: r.grade for r in submission.uace_results.all()}
                uce_grades = [g.grade for g in submission.uce_grades.all()]
                if uace_results:
                    weight = calculate_weight(
                        student_results=uace_results,
                        course_requirements=[],
                        uce_grades=uce_grades,
                        gender=submission.gender,
                    )
                    weights.append(weight)
            if weights:
                avg_weight = sum(weights) / len(weights)
        
        return Response(
            {
                "top_courses": top_courses_data,
                "common_profiles": common_profiles,
                "total_submissions": total_submissions,
                "total_candidates": total_candidates,
                "average_weight": str(avg_weight),
            },
            status=status.HTTP_200_OK,
        )


class CourseSubjectRequirementListAPIView(APIView):
    """Get course subject requirements for admin management."""
    
    @staticmethod
    def _require_admin_token(request):
        """Verify admin token."""
        auth_header = request.META.get("HTTP_AUTHORIZATION", "")
        if not auth_header.startswith("Bearer "):
            raise AuthenticationFailed("admin token required")
        
        token = auth_header[7:]
        try:
            session = AdminSessionToken.objects.get(token=token)
            session.last_used_at = timezone.now()
            session.save()
            return session.user
        except AdminSessionToken.DoesNotExist:
            raise AuthenticationFailed("invalid token")
    
    def get(self, request, course_id=None, *args, **kwargs):
        self._require_admin_token(request)
        
        course_id = course_id or request.query_params.get("course_id")
        if not course_id:
            raise PermissionDenied("course_id required")
        
        try:
            course = Course.objects.get(id=course_id)
        except Course.DoesNotExist:
            raise PermissionDenied("Course not found")
        
        requirements = CourseSubjectRequirement.objects.filter(course=course).select_related("subject")
        data = [
            {
                "id": req.id,
                "subject_name": req.subject.name,
                "subject_id": req.subject.id,
                "category": req.category,
            }
            for req in requirements
        ]
        return Response(data, status=status.HTTP_200_OK)


class CourseSubjectRequirementUpdateAPIView(APIView):
    """Add, update, or delete course subject requirements."""
    
    @staticmethod
    def _require_admin_token(request):
        """Verify admin token."""
        auth_header = request.META.get("HTTP_AUTHORIZATION", "")
        if not auth_header.startswith("Bearer "):
            raise AuthenticationFailed("admin token required")
        
        token = auth_header[7:]
        try:
            session = AdminSessionToken.objects.get(token=token)
            session.last_used_at = timezone.now()
            session.save()
            return session.user
        except AdminSessionToken.DoesNotExist:
            raise AuthenticationFailed("invalid token")
    
    def post(self, request, *args, **kwargs):
        """Add a new subject requirement for a course."""
        self._require_admin_token(request)
        
        course_id = request.data.get("course_id")
        subject_id = request.data.get("subject_id")
        category = request.data.get("category")
        
        if not all([course_id, subject_id, category]):
            raise PermissionDenied("course_id, subject_id, and category required")
        
        if category not in ["essential", "relevant", "desirable"]:
            raise PermissionDenied("Invalid category")
        
        try:
            course = Course.objects.get(id=course_id)
            subject = Subject.objects.get(id=subject_id)
        except (Course.DoesNotExist, Subject.DoesNotExist):
            raise PermissionDenied("Course or subject not found")
        
        req, created = CourseSubjectRequirement.objects.get_or_create(
            course=course,
            subject=subject,
            defaults={"category": category}
        )
        
        if not created:
            req.category = category
            req.save()
        
        return Response(
            {
                "id": req.id,
                "subject_id": req.subject.id,
                "subject_name": req.subject.name,
                "category": req.category,
            },
            status=status.HTTP_201_CREATED if created else status.HTTP_200_OK,
        )
    
    def delete(self, request, *args, **kwargs):
        """Delete a subject requirement."""
        self._require_admin_token(request)
        
        requirement_id = request.query_params.get("requirement_id")
        if not requirement_id:
            raise PermissionDenied("requirement_id required")
        
        try:
            req = CourseSubjectRequirement.objects.get(id=requirement_id)
            req.delete()
            return Response({"message": "Deleted"}, status=status.HTTP_204_NO_CONTENT)
        except CourseSubjectRequirement.DoesNotExist:
            raise PermissionDenied("Requirement not found")
